"""Extract enrolment trends over time for a set of schools from the historical ACARA data.

Usage:
  python acara_trends.py <xlsx_path> --schools schools.json [--output output.json]

Reads the 2008-2025 historical file, filters to the given schools, and produces
year-over-year enrolment totals and growth rates.
"""
import argparse
import json

def main():
    parser = argparse.ArgumentParser(description="ACARA enrolment trends")
    parser.add_argument("xlsx_path", help="Path to historical ACARA XLSX (2008-2025)")
    parser.add_argument("--schools", help="JSON file with array of school names", required=True)
    parser.add_argument("--output", help="Output JSON file path")
    args = parser.parse_args()

    import openpyxl

    with open(args.schools, encoding="utf-8") as f:
        target_names = [n.strip().lower() for n in json.load(f)]

    wb = openpyxl.load_workbook(args.xlsx_path, read_only=True)
    sheet_name = [s for s in wb.sheetnames if s != "DataDictionary"][0]
    ws = wb[sheet_name]

    rows = ws.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows)]

    year_idx = headers.index("Calendar Year")
    name_idx = headers.index("School Name")
    enrol_idx = headers.index("Total Enrolments")

    yearly_totals = {}

    for row in rows:
        name = (row[name_idx] or "").strip().lower()
        matched = any(t in name or name in t for t in target_names)
        if not matched:
            continue

        year = row[year_idx]
        enrol = row[enrol_idx] or 0
        if year not in yearly_totals:
            yearly_totals[year] = {"total_enrolments": 0, "school_count": 0}
        yearly_totals[year]["total_enrolments"] += enrol
        yearly_totals[year]["school_count"] += 1

    wb.close()

    # Sort by year and compute growth
    sorted_years = sorted(yearly_totals.keys())
    trends = []
    for i, year in enumerate(sorted_years):
        entry = {
            "year": year,
            "total_enrolments": yearly_totals[year]["total_enrolments"],
            "school_count": yearly_totals[year]["school_count"],
        }
        if i > 0:
            prev = yearly_totals[sorted_years[i - 1]]["total_enrolments"]
            curr = entry["total_enrolments"]
            entry["yoy_growth_pct"] = round((curr - prev) / prev * 100, 1) if prev else None
        trends.append(entry)

    result = {
        "school_names_matched": len(target_names),
        "years_covered": len(trends),
        "trends": trends,
    }

    output = json.dumps(result, ensure_ascii=False, indent=2)
    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(output)
        print(f"Trends for {len(trends)} years written to {args.output}")
    else:
        print(output)

if __name__ == "__main__":
    main()
