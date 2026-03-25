"""Filter ACARA school profile data for a specific set of schools.

Usage:
  python filter_acara.py <xlsx_path> --schools schools.json [--output output.json]
  python filter_acara.py <xlsx_path> --match-names "School A,School B" [--output output.json]
  python filter_acara.py <xlsx_path> --state NSW --sector Catholic [--output output.json]

schools.json should be a JSON array of school names to match against.
Output: JSON with matched school records and aggregate stats.
"""
import argparse
import json
import sys

def load_workbook_data(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    # Find the data sheet (not DataDictionary)
    sheet_name = [s for s in wb.sheetnames if s != "DataDictionary"][0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    data = []
    for row in rows[1:]:
        record = dict(zip(headers, row))
        data.append(record)
    wb.close()
    return headers, data

def filter_schools(data, school_names=None, state=None, sector=None):
    results = []
    for record in data:
        if state and record.get("State") != state:
            continue
        if sector and record.get("School Sector") != sector:
            continue
        if school_names:
            name = (record.get("School Name") or "").strip()
            # Fuzzy match: check if any target name is contained in or matches the school name
            matched = False
            for target in school_names:
                target_clean = target.strip().lower()
                name_clean = name.lower()
                if target_clean in name_clean or name_clean in target_clean:
                    matched = True
                    break
            if not matched:
                continue
        results.append(record)
    return results

def compute_aggregates(records):
    enrolments = [r.get("Total Enrolments") or 0 for r in records]
    fte_teaching = [r.get("Full Time Equivalent Teaching Staff") or 0 for r in records]
    icseas = [r.get("ICSEA") for r in records if r.get("ICSEA")]
    indigenous = [r.get("Indigenous Enrolments (%)") for r in records if r.get("Indigenous Enrolments (%)") is not None]
    lbote = [r.get("Language Background Other Than English - Yes (%)") for r in records if r.get("Language Background Other Than English - Yes (%)") is not None]

    total_enrolment = sum(enrolments)
    total_fte = sum(fte_teaching)

    return {
        "school_count": len(records),
        "total_enrolments": total_enrolment,
        "avg_enrolments_per_school": round(total_enrolment / len(records), 0) if records else 0,
        "total_fte_teaching": round(total_fte, 1),
        "student_teacher_ratio": round(total_enrolment / total_fte, 1) if total_fte else None,
        "icsea_min": min(icseas) if icseas else None,
        "icsea_max": max(icseas) if icseas else None,
        "icsea_avg": round(sum(icseas) / len(icseas), 0) if icseas else None,
        "avg_indigenous_pct": round(sum(indigenous) / len(indigenous), 1) if indigenous else None,
        "avg_lbote_pct": round(sum(lbote) / len(lbote), 1) if lbote else None,
    }

def main():
    parser = argparse.ArgumentParser(description="Filter ACARA data")
    parser.add_argument("xlsx_path", help="Path to ACARA XLSX file")
    parser.add_argument("--schools", help="JSON file with array of school names to match")
    parser.add_argument("--match-names", help="Comma-separated school names to match")
    parser.add_argument("--state", help="Filter by state (e.g. NSW)")
    parser.add_argument("--sector", help="Filter by sector (e.g. Catholic)")
    parser.add_argument("--output", help="Output JSON file path")
    args = parser.parse_args()

    headers, data = load_workbook_data(args.xlsx_path)

    school_names = None
    if args.schools:
        with open(args.schools, encoding="utf-8") as f:
            school_names = json.load(f)
    elif args.match_names:
        school_names = [n.strip() for n in args.match_names.split(",")]

    filtered = filter_schools(data, school_names, args.state, args.sector)
    aggregates = compute_aggregates(filtered)

    result = {
        "filter": {
            "state": args.state,
            "sector": args.sector,
            "school_names_provided": len(school_names) if school_names else 0,
        },
        "aggregates": aggregates,
        "schools": [
            {
                "name": r.get("School Name"),
                "suburb": r.get("Suburb"),
                "postcode": r.get("Postcode"),
                "type": r.get("School Type"),
                "year_range": r.get("Year Range"),
                "total_enrolments": r.get("Total Enrolments"),
                "fte_teaching": r.get("Full Time Equivalent Teaching Staff"),
                "icsea": r.get("ICSEA"),
                "indigenous_pct": r.get("Indigenous Enrolments (%)"),
                "lbote_pct": r.get("Language Background Other Than English - Yes (%)"),
                "geolocation": r.get("Geolocation"),
                "acara_id": r.get("ACARA SML ID"),
            }
            for r in filtered
        ],
    }

    output = json.dumps(result, ensure_ascii=False, indent=2)
    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(output)
        print(f"Filtered {len(filtered)} schools, aggregates computed. Written to {args.output}")
    else:
        print(output)

if __name__ == "__main__":
    main()
